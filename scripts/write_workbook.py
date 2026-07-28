"""Shared xlsx writer for finance-superpowers deliverables.

Every deliverable workbook is produced through this module so the
formatting bar lives in one tested place: cover sheet with sources and
assumptions, frozen headers, accounting number formats, styled totals.
"""
import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MONEY_FMT = "#,##0.00;(#,##0.00)"
MAX_SHEET_NAME_LEN = 31
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)
TOP_BORDER = Border(top=Side(style="thin"))


def _parse_money(val):
    """Parse a money string into a float.

    Handles thousands separators (1,234.00), a leading currency symbol
    ($5,000.00), and accounting-style parenthesized negatives
    ((1,234.00) -> -1234.00), including both together (($6.50)).
    """
    s = str(val).strip()
    negative = s.startswith("(") and s.endswith(")")
    if negative:
        s = s[1:-1].strip()
    s = s.replace(",", "").replace("$", "").strip()
    num = float(s)
    return -num if negative else num


def _autofit(ws):
    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)


def _write_cover(wb, cover):
    ws = wb.active
    ws.title = "cover"
    ws["A1"] = cover["title"]
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"], ws["B3"] = "Engagement:", cover.get("engagement", "")
    ws["A4"], ws["B4"] = "Date:", cover.get("date", "")
    row = 6
    ws.cell(row=row, column=1, value="Sources").font = Font(bold=True)
    for src in cover.get("sources", []):
        row += 1
        ws.cell(row=row, column=1, value=src)
    row += 2
    ws.cell(row=row, column=1, value="Assumptions").font = Font(bold=True)
    assumptions = cover.get("assumptions", []) or ["None."]
    for item in assumptions:
        row += 1
        ws.cell(row=row, column=1, value=item)
    _autofit(ws)


def _write_sheet(wb, spec):
    name = spec["name"]
    if len(name) > MAX_SHEET_NAME_LEN:
        name = name[:MAX_SHEET_NAME_LEN]
    ws = wb.create_sheet(name)
    with open(spec["csv"], newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        # empty source csv: nothing to write, but don't crash the run.
        ws.freeze_panes = "A2"
        _autofit(ws)
        return
    header, data = rows[0], rows[1:]
    money_cols = spec.get("money_cols", [])
    unmatched = [c for c in money_cols if c not in header]
    if unmatched:
        raise ValueError(
            f"money_cols {unmatched!r} not found in header {header!r} of "
            f"sheet {spec['name']!r} ({spec['csv']}) -- check for a case "
            f"or spelling mismatch; a silently-skipped money column would "
            f"write numbers as unsummable text with no warning."
        )
    money_idx = {header.index(c) for c in money_cols}
    for j, colname in enumerate(header, start=1):
        cell = ws.cell(row=1, column=j, value=colname)
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL
    for i, row in enumerate(data, start=2):
        for j, val in enumerate(row, start=1):
            if (j - 1) in money_idx and val not in ("", None):
                cell = ws.cell(row=i, column=j, value=_parse_money(val))
                cell.number_format = MONEY_FMT
            else:
                ws.cell(row=i, column=j, value=val)
    for r in spec.get("total_rows", []):
        for j in range(1, len(header) + 1):
            cell = ws.cell(row=r + 2, column=j)
            cell.font, cell.border = TOTAL_FONT, TOP_BORDER
    ws.freeze_panes = "A2"
    _autofit(ws)


def write_workbook(output_path, cover, sheets):
    wb = Workbook()
    _write_cover(wb, cover)
    for spec in sheets:
        _write_sheet(wb, spec)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main():
    manifest = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    out = write_workbook(manifest["output"], manifest["cover"], manifest["sheets"])
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
