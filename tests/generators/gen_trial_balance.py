"""Generate the Brightwater ragged trial balance fixture.

Economic model
--------------
This is a six-months-ended June 30, 2026 interim consolidated trial balance
for the *same* fictional company whose FY2025 diligence data room
(sample-data/brightwater/data-room/, built by gen_data_room.py) reports:
  revenue 49,200,000 gross, COGS 36,210,000, opex 9,180,000,
  EBITDA 3,560,000, pre-tax income 2,420,000; and a 2025-12-31 balance
  sheet of cash 2,140,000 / AR net 4,610,000 / inventory 5,890,000 /
  fixed assets net 3,220,000 / AP 3,410,000 / accrued 1,120,000 /
  term loan 4,800,000 / equity 6,530,000.

This TB is a *different period* (interim, six months later) and is not
required to tie to those figures exactly, but it must be recognizably the
same business at roughly half-year scale: consolidated revenue here is
~25.6M (annualizing to ~51M, in line with the data room's 49.2M), total
expenses are less than revenue (the company is profitable, consistent
with the data room's positive pre-tax income), and the term loan balance
(4,500,000 = 300,000 current + 4,200,000 long-term, all held at holdco)
matches the 2026-06-30 balance on the data room's amortization schedule
(4,800,000 less two quarterly principal payments of 150,000).

Entity shape:
  - brightwater us: the large operating entity (full P&L and balance sheet).
  - brightwater canada: a smaller operating entity, same account structure.
  - brightwater holdco: the parent, no operating revenue or expenses other
    than interest on the term loan it holds; no AR/inventory/fixed assets
    since it has no operations. Its retained earnings is the single
    consolidation plug that zeroes the whole trial balance.
  - Contra accounts (allowance for doubtful accounts, accumulated
    depreciation) are sized smaller in magnitude than the asset they
    offset, in every entity that carries them.

Self-checking: asserts the clean account rows net to exactly zero, the
planted hazards are present, and the plausibility invariants above before
saving. Deterministic -- no randomness; all amounts are fixed literals.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUT = Path(__file__).resolve().parents[2] / "sample-data" / "brightwater" / "trial-balance" / "tb_2026-06.xlsx"

# Reference point only (not required to tie exactly -- different period).
DATA_ROOM_ANNUAL_REVENUE = 49_200_000

SECTIONS = [
    ("assets", [(1000, "cash - operating"), (1010, "cash - payroll"),
                (1200, "accounts receivable"), (1210, "allowance for doubtful accounts"),
                (1300, "inventory - finished goods"), (1400, "prepaid expenses"),
                (1500, "fixed assets"), (1510, "accumulated depreciation")]),
    ("liabilities", [(2000, "accounts payable"), (2100, "accrued liabilities"),
                     (2200, "deferred revenue"), (2500, "term loan - current"),
                     (2600, "term loan - long term")]),
    ("equity", [(3000, "common stock"), (3900, "retained earnings")]),
    ("revenue", [(4000, "product revenue"), (4100, "freight revenue")]),
    ("expenses", [(5000, "cost of goods sold"), (5100, "freight out"),
                  (5210, "salaries and wages"), (5220, "employee benefits"),
                  (5300, "rent"), (5400, "depreciation"), (5900, "interest expense")]),
]
ENTITIES = ["brightwater us", "brightwater canada", "brightwater holdco"]

# Explicit, hand-set, economically plausible amounts per entity/account.
# Sign convention (standard TB signed-debit representation):
#   assets +, contra-assets -, liabilities -, equity -, revenue -, expenses +.
# 3900 (retained earnings) is deliberately absent here -- it is computed as
# the single balancing plug on holdco in build_rows(), as before.
#
# brightwater canada / 5210 = 58,349.10 and brightwater us / 1400 = 77,614.00
# are pinned to their exact values: skills/clean/SKILL.md and
# skills/understand/SKILL.md quote them verbatim, and the duplicate-row /
# text-amount hazards are planted on exactly these two rows below.
AMOUNTS = {
    "brightwater us": {
        1000: 1_380_000.00, 1010: 165_000.00,
        1200: 3_050_000.00, 1210: -138_000.00,
        1300: 3_480_000.00, 1400: 77_614.00,
        1500: 2_760_000.00, 1510: -1_120_000.00,
        2000: -1_940_000.00, 2100: -560_000.00, 2200: -190_000.00,
        2500: 0.00, 2600: 0.00,
        3000: -500_000.00,
        4000: -22_300_000.00, 4100: -600_000.00,
        5000: 17_400_000.00, 5100: 410_000.00, 5210: 3_550_000.00,
        5220: 560_000.00, 5300: 145_000.00, 5400: 165_000.00, 5900: 45_000.00,
    },
    "brightwater canada": {
        1000: 320_000.00, 1010: 45_000.00,
        1200: 780_000.00, 1210: -32_000.00,
        1300: 890_000.00, 1400: 28_500.00,
        1500: 640_000.00, 1510: -240_000.00,
        2000: -410_000.00, 2100: -125_000.00, 2200: -48_000.00,
        2500: 0.00, 2600: 0.00,
        3000: -150_000.00,
        4000: -2_650_000.00, 4100: -95_000.00,
        5000: 2_050_000.00, 5100: 62_000.00, 5210: 58_349.10,
        5220: 74_000.00, 5300: 38_000.00, 5400: 42_000.00, 5900: 12_000.00,
    },
    "brightwater holdco": {
        # No operations: no AR/inventory/fixed assets/deferred revenue/
        # revenue/COGS/opex. Only cash, a small prepaid, minor payables,
        # the consolidated term loan, common stock, and loan interest.
        1000: 209_999.89, 1400: 8_500.00,
        2000: -15_000.00, 2100: -25_000.00,
        2500: -300_000.00, 2600: -4_200_000.00,
        3000: -1_500_000.00,
        5900: 95_000.00,
    },
}


def amount(entity, acct):
    return AMOUNTS[entity].get(acct)


def build_rows():
    rows, balance = [], 0.0
    for entity in ENTITIES:
        rows.append(("entity_header", entity, None, None))
        for section, accounts in SECTIONS:
            subtotal = 0.0
            for acct, name in accounts:
                if (acct, name) == (3900, "retained earnings"):
                    continue  # placeholder; balanced below
                amt = amount(entity, acct)
                if amt is None:
                    continue  # account not applicable to this entity
                rows.append(("account", entity, acct, (name, amt)))
                subtotal += amt
                balance += amt
            rows.append(("subtotal", entity, None, (f"total {section}", round(subtotal, 2))))
    # single consolidated retained-earnings balancer on holdco
    rows.append(("account", "brightwater holdco", 3900,
                 ("retained earnings", round(-balance, 2))))
    return rows


def main():
    rows = build_rows()
    clean_sum = round(sum(r[3][1] for r in rows if r[0] == "account"), 2)
    if clean_sum == 0.0:
        clean_sum = 0.0  # normalize -0.0 (floating-point residue) to 0.0
    assert clean_sum == 0.0, f"TB does not balance: {clean_sum}"

    # --- plausibility invariants -------------------------------------
    def total(acct):
        return sum(a for e, ac, a in
                   ((r[1], r[2], r[3][1]) for r in rows if r[0] == "account")
                   if ac == acct)

    gross_ar = total(1200)
    allowance = total(1210)
    assert 0 < abs(allowance) < abs(gross_ar), (
        f"allowance {allowance} must be smaller in magnitude than gross AR {gross_ar}")

    fixed_assets = total(1500)
    accum_dep = total(1510)
    assert 0 < abs(accum_dep) < abs(fixed_assets), (
        f"accumulated depreciation {accum_dep} must be smaller in magnitude "
        f"than fixed assets {fixed_assets}")

    revenue = total(4000) + total(4100)
    assert revenue < 0, f"revenue should be a credit balance, got {revenue}"
    annualized_revenue = abs(revenue) * 2
    # half-year revenue should annualize to roughly the data room's FY2025
    # gross revenue (same company, different period) -- within +/-25%.
    assert 0.75 * DATA_ROOM_ANNUAL_REVENUE < annualized_revenue < 1.25 * DATA_ROOM_ANNUAL_REVENUE, (
        f"annualized revenue {annualized_revenue} not consistent with "
        f"data room revenue {DATA_ROOM_ANNUAL_REVENUE}")

    expense_accts = [5000, 5100, 5210, 5220, 5300, 5400, 5900]
    total_expenses = sum(total(a) for a in expense_accts)
    assert total_expenses < abs(revenue), (
        f"expenses {total_expenses} must be less than revenue {abs(revenue)} "
        f"(the company is profitable, per the data room's pre-tax income)")

    wb = Workbook()
    ws = wb.active
    ws.title = "tb"
    ws.merge_cells("A1:E2")
    ws["A1"] = "Brightwater Distribution Co. — Consolidated Trial Balance — June 30, 2026 (all entities, local currency USD)"
    ws["A1"].font = Font(bold=True)
    ws.append([])
    ws.append(["entity", "account", "description", "", "balance"])
    dup_written = False
    text_amount_written = False
    for kind, entity, acct, payload in rows:
        if kind == "entity_header":
            ws.append([entity.upper(), None, None, None, None])
        elif kind == "subtotal":
            name, amt = payload
            ws.append([None, None, name.upper(), None, amt])
        else:
            name, amt = payload
            if acct == 1400 and not text_amount_written:
                amt_out = f"{amt:,.2f}"  # planted: amount stored as text
                text_amount_written = True
            else:
                amt_out = amt
            ws.append([entity, acct, name, None, amt_out])
            if acct == 5210 and entity == "brightwater canada" and not dup_written:
                ws.append([entity, acct, name, None, amt])  # planted duplicate
                dup_written = True
    grand = round(sum(r[3][1] for r in rows if r[0] != "entity_header"), 2)
    ws.append([None, None, "GRAND TOTAL", None, grand])

    assert dup_written and text_amount_written
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}; clean sum {clean_sum}; naive grand total {grand}")
    print(f"gross AR {gross_ar:,.2f} vs allowance {allowance:,.2f}")
    print(f"fixed assets {fixed_assets:,.2f} vs accumulated depreciation {accum_dep:,.2f}")
    print(f"revenue {revenue:,.2f} vs total expenses {total_expenses:,.2f}")


if __name__ == "__main__":
    main()
