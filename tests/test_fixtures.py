"""Guards on the committed Brightwater fixtures: the trial balance, its
tidy derivative, and the data-room extracts.

These assert against the files as committed under sample-data/ and
tests/fixtures/ -- they do NOT regenerate anything. The point is that an
accidental hand-edit to a fixture (or a generator change that isn't
re-run everywhere) is caught by `python -m pytest` without anyone needing
to know to run tests/generators/*.py by hand.
"""
import csv
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TB_PATH = ROOT / "sample-data" / "brightwater" / "trial-balance" / "tb_2026-06.xlsx"
TB_TIDY_PATH = ROOT / "tests" / "fixtures" / "tb_tidy.csv"
DATA_ROOM_FINANCIALS = ROOT / "sample-data" / "brightwater" / "data-room" / "financials"

DATA_ROOM_ANNUAL_REVENUE = 49_200_000
DATA_ROOM_FY2025_CLOSING_EQUITY = 6_530_000
COGS_TOTAL = 36_210_000
OPEX_TOTAL = 9_180_000
INTEREST_TOTAL = 1_140_000

ZERO_TOL = 0.005  # footing tolerance: sub-cent float residue only


def _as_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v).replace(",", "").strip())


def _load_tb():
    wb = load_workbook(TB_PATH)
    ws = wb.active
    return ws, list(ws.iter_rows(values_only=True))


def _account_rows(rows):
    """Individual account postings only -- excludes the title, blank row,
    header row, entity-caption rows, subtotal rows, and the grand-total row."""
    out = []
    for r in rows:
        entity, acct, desc, _, balance = r
        if isinstance(entity, str) and entity.startswith("brightwater") and isinstance(acct, int):
            out.append((entity, acct, desc, _as_number(balance)))
    return out


def _dedupe(accounts):
    """Drop exact-duplicate rows (same entity/account/description/balance),
    keeping the first occurrence -- what a clean step's own footing would
    do with the planted duplicate. Used for every economic assertion below;
    the raw (non-deduped) rows are used only to confirm the hazard exists."""
    seen = set()
    out = []
    for a in accounts:
        if a in seen:
            continue
        seen.add(a)
        out.append(a)
    return out


def _total(accounts, acct):
    return sum(a[3] for a in accounts if a[1] == acct)


def test_tb_clean_rows_foot_to_zero():
    _, rows = _load_tb()
    accounts = _dedupe(_account_rows(rows))
    total = sum(a[3] for a in accounts)
    assert abs(total) < ZERO_TOL, f"TB account rows (deduped) do not foot to 0.00: {total}"


def test_tb_hazards_present():
    ws, rows = _load_tb()
    assert "A1:E2" in [str(r) for r in ws.merged_cells.ranges], "merged title A1:E2 missing"

    entity_headers = [r for r in rows if r[0] and str(r[0]).isupper()
                       and str(r[0]).startswith("BRIGHTWATER")]
    assert len(entity_headers) == 3, f"expected 3 entity header rows, found {len(entity_headers)}"

    subtotal_rows = [r for r in rows if r[2] and str(r[2]).startswith("TOTAL")]
    assert len(subtotal_rows) == 15, f"expected 15 subtotal/caption rows, found {len(subtotal_rows)}"

    dup_rows = [r for r in rows if r[1] == 5210 and r[0] == "brightwater canada"]
    assert len(dup_rows) == 2, f"expected canada 5210 duplicated once (2 rows), found {len(dup_rows)}"

    text_rows = [r for r in rows if r[1] == 1400 and isinstance(r[4], str)]
    assert len(text_rows) == 1, f"expected exactly 1 text-typed amount, found {len(text_rows)}"

    grand_rows = [r for r in rows if r[2] == "GRAND TOTAL"]
    assert len(grand_rows) == 1, f"expected exactly 1 grand-total row, found {len(grand_rows)}"


def test_pinned_amounts_exact():
    ws, rows = _load_tb()
    dup_rows = [r for r in rows if r[1] == 5210 and r[0] == "brightwater canada"]
    assert dup_rows, "canada 5210 row not found"
    assert all(_as_number(r[4]) == 58349.10 for r in dup_rows), (
        "canada 5210 duplicate amount must be exactly 58,349.10")

    text_rows = [r for r in rows if r[1] == 1400 and isinstance(r[4], str)]
    assert text_rows, "no text-typed 1400 row found"
    assert text_rows[0][0] == "brightwater us", "text-typed amount must be on brightwater us"
    assert text_rows[0][4] == "77,614.00", (
        f"us 1400 text amount must be exactly '77,614.00', got {text_rows[0][4]!r}")


def test_plausibility_invariants():
    _, rows = _load_tb()
    accounts = _dedupe(_account_rows(rows))

    gross_ar, allowance = _total(accounts, 1200), _total(accounts, 1210)
    assert 0 < abs(allowance) < abs(gross_ar), (
        f"allowance {allowance} must be smaller in magnitude than gross AR {gross_ar}")

    fixed_assets, accum_dep = _total(accounts, 1500), _total(accounts, 1510)
    assert 0 < abs(accum_dep) < abs(fixed_assets), (
        f"accumulated depreciation {accum_dep} must be smaller in magnitude "
        f"than fixed assets {fixed_assets}")

    revenue = _total(accounts, 4000) + _total(accounts, 4100)
    assert revenue < 0, "revenue should be a credit balance"
    annualized_revenue = abs(revenue) * 2
    assert 0.75 * DATA_ROOM_ANNUAL_REVENUE < annualized_revenue < 1.25 * DATA_ROOM_ANNUAL_REVENUE, (
        f"annualized revenue {annualized_revenue} not consistent with the "
        f"data room's {DATA_ROOM_ANNUAL_REVENUE}")

    expense_accts = (5000, 5100, 5210, 5220, 5300, 5400, 5900)
    total_expenses = sum(_total(accounts, a) for a in expense_accts)
    assert total_expenses < abs(revenue), "expenses must be less than revenue (profitable)"

    # per-entity, for the two operating entities (holdco has no operations)
    for entity in ("brightwater us", "brightwater canada"):
        entity_accounts = [a for a in accounts if a[0] == entity]
        ar_e = _total(entity_accounts, 1200)
        allw_e = _total(entity_accounts, 1210)
        assert 0 < abs(allw_e) < abs(ar_e), f"{entity}: allowance must be smaller than gross AR"
        fa_e = _total(entity_accounts, 1500)
        ad_e = _total(entity_accounts, 1510)
        assert 0 < abs(ad_e) < abs(fa_e), f"{entity}: accum dep must be smaller than fixed assets"


def test_equity_rolls_forward_from_data_room_close():
    """Retained earnings must continue the data room's FY2025 close, not
    be an arbitrary balancing plug: opening RE (2025-12-31) = FY2025
    closing equity minus common stock, and closing RE (this TB) = opening
    RE plus this period's net income."""
    _, rows = _load_tb()
    accounts = _dedupe(_account_rows(rows))

    revenue = _total(accounts, 4000) + _total(accounts, 4100)
    expense_accts = (5000, 5100, 5210, 5220, 5300, 5400, 5900)
    total_expenses = sum(_total(accounts, a) for a in expense_accts)
    net_income = -revenue - total_expenses

    common_stock = _total(accounts, 3000)
    retained_earnings = _total(accounts, 3900)
    opening_retained_earnings = DATA_ROOM_FY2025_CLOSING_EQUITY - abs(common_stock)
    expected_re = round(-(opening_retained_earnings + net_income), 2)
    assert round(retained_earnings, 2) == expected_re, (
        f"retained earnings {retained_earnings} does not roll forward from "
        f"opening {opening_retained_earnings} + net income {net_income}")

    consolidated_equity = round(abs(common_stock) + abs(retained_earnings), 2)
    expected_equity = round(DATA_ROOM_FY2025_CLOSING_EQUITY + net_income, 2)
    assert consolidated_equity == expected_equity, (
        f"consolidated equity {consolidated_equity} is not continuous with "
        f"the data room's FY2025 closing equity {DATA_ROOM_FY2025_CLOSING_EQUITY} "
        f"plus this period's net income {net_income}")


def test_tb_tidy_foots_to_zero_no_hazards():
    with open(TB_TIDY_PATH, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert rows, "tb_tidy.csv is empty"

    total = sum(float(r["balance"]) for r in rows)
    assert abs(total) < ZERO_TOL, f"tb_tidy.csv does not foot to 0.00: {total}"

    # no subtotal/caption rows -- every row's account is a plain account number
    assert all(r["account"].strip().isdigit() for r in rows), (
        "tb_tidy.csv must contain only account rows, no subtotal/caption rows")

    # no duplicate rows (the xlsx's planted duplicate must have been resolved)
    keys = [(r["entity"], r["account"], r["description"], r["balance"]) for r in rows]
    assert len(keys) == len(set(keys)), "tb_tidy.csv has duplicate rows"

    canada_5210 = [r for r in rows if r["entity"] == "brightwater canada" and r["account"] == "5210"]
    assert len(canada_5210) == 1 and canada_5210[0]["balance"] == "58349.10", (
        "tb_tidy.csv canada 5210 must appear exactly once at 58349.10")

    us_1400 = [r for r in rows if r["entity"] == "brightwater us" and r["account"] == "1400"]
    assert len(us_1400) == 1 and us_1400[0]["balance"] == "77614.00", (
        "tb_tidy.csv us 1400 must appear exactly once, coerced to 77614.00")


def _read_tb_extract(entity):
    path = DATA_ROOM_FINANCIALS / f"tb_extract_{entity}_fy2025.csv"
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def test_data_room_revenue_identity():
    total = 0
    for entity in ("us", "canada", "holdco"):
        for row in _read_tb_extract(entity):
            if int(row["account"]) in (4000, 4100):
                total += int(row["fy2025_amount"])
    assert total == DATA_ROOM_ANNUAL_REVENUE


def test_data_room_expense_consolidation():
    cogs = opex = interest = 0
    for entity in ("us", "canada", "holdco"):
        for row in _read_tb_extract(entity):
            acct = int(row["account"])
            amt = int(row["fy2025_amount"])
            if acct == 5000:
                cogs += amt
            elif acct in (5100, 5210, 5220):
                opex += amt
            elif acct == 5900:
                interest += amt
    assert cogs == COGS_TOTAL, f"cogs {cogs} != {COGS_TOTAL}"
    assert opex == OPEX_TOTAL, f"opex {opex} != {OPEX_TOTAL}"
    assert interest == INTEREST_TOTAL, f"interest {interest} != {INTEREST_TOTAL}"
