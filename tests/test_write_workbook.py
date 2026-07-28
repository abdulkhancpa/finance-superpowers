import csv
import json
import subprocess
from pathlib import Path

from openpyxl import load_workbook

from scripts.write_workbook import write_workbook, MONEY_FMT


def make_csv(tmp_path: Path) -> Path:
    p = tmp_path / "sched.csv"
    with open(p, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["account", "description", "amount"])
        w.writerow(["1000", "cash", "1234.5"])
        w.writerow(["1200", "ar", "-200"])
        w.writerow(["", "total", "1034.5"])
    return p


def build(tmp_path: Path) -> Path:
    out = tmp_path / "out" / "deliverable.xlsx"
    return write_workbook(
        out,
        cover={
            "title": "test deliverable",
            "engagement": "unit test",
            "date": "2026-07-27",
            "sources": ["sources/tb.xlsx"],
            "assumptions": ["threshold: $1,000"],
        },
        sheets=[{"name": "schedule", "csv": make_csv(tmp_path),
                 "money_cols": ["amount"], "total_rows": [2]}],
    )


def test_creates_cover_and_sheet(tmp_path):
    wb = load_workbook(build(tmp_path))
    assert wb.sheetnames == ["cover", "schedule"]
    cover = wb["cover"]
    assert cover["A1"].value == "test deliverable"
    texts = [c.value for row in cover.iter_rows() for c in row if c.value]
    assert "sources/tb.xlsx" in texts and "threshold: $1,000" in texts


def test_sheet_formatting(tmp_path):
    ws = load_workbook(build(tmp_path))["schedule"]
    assert ws.freeze_panes == "A2"
    assert ws["A1"].font.bold and ws["A1"].fill.fgColor.rgb.endswith("1F3864")
    assert ws["C2"].value == 1234.5 and ws["C2"].number_format == MONEY_FMT
    assert ws["B4"].font.bold  # total row styled


def test_empty_assumptions_says_none(tmp_path):
    out = write_workbook(tmp_path / "d.xlsx", cover={"title": "t"}, sheets=[])
    cover = load_workbook(out)["cover"]
    texts = [c.value for row in cover.iter_rows() for c in row if c.value]
    assert "None." in texts


def test_accounting_negatives_and_currency_symbols(tmp_path):
    p = tmp_path / "acct.csv"
    with open(p, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["account", "description", "amount"])
        w.writerow(["1000", "cash", "$5,000.00"])
        w.writerow(["1200", "ar", "(1,234.00)"])
        w.writerow(["1300", "other", "($6.50)"])
    out = write_workbook(
        tmp_path / "out" / "acct.xlsx",
        cover={"title": "t"},
        sheets=[{"name": "schedule", "csv": p, "money_cols": ["amount"], "total_rows": []}],
    )
    ws = load_workbook(out)["schedule"]
    assert ws["C2"].value == 5000.0
    assert ws["C3"].value == -1234.0
    assert ws["C4"].value == -6.5


def test_money_cols_mismatch_raises(tmp_path):
    p = make_csv(tmp_path)
    import pytest
    with pytest.raises(ValueError):
        write_workbook(
            tmp_path / "out" / "bad.xlsx",
            cover={"title": "t"},
            # "Amount" (capital A) does not match the "amount" header
            sheets=[{"name": "schedule", "csv": p, "money_cols": ["Amount"], "total_rows": []}],
        )


def test_empty_csv_does_not_crash(tmp_path):
    p = tmp_path / "empty.csv"
    p.write_text("", encoding="utf-8")
    out = write_workbook(
        tmp_path / "out" / "empty.xlsx",
        cover={"title": "t"},
        sheets=[{"name": "schedule", "csv": p, "money_cols": [], "total_rows": []}],
    )
    wb = load_workbook(out)
    assert "schedule" in wb.sheetnames


def test_long_sheet_name_truncated_no_warning(tmp_path, recwarn):
    p = make_csv(tmp_path)
    long_name = "a" * 40
    out = write_workbook(
        tmp_path / "out" / "long.xlsx",
        cover={"title": "t"},
        sheets=[{"name": long_name, "csv": p, "money_cols": ["amount"], "total_rows": []}],
    )
    wb = load_workbook(out)
    assert all(len(name) <= 31 for name in wb.sheetnames)
    assert not any("more than 31 characters" in str(w.message) for w in recwarn.list)


def test_cli_manifest(tmp_path):
    csv_path = make_csv(tmp_path)
    out_path = tmp_path / "cli_output.xlsx"
    manifest = {
        "output": str(out_path),
        "cover": {
            "title": "CLI test",
            "engagement": "manifest test",
            "date": "2026-07-27",
        },
        "sheets": [
            {
                "name": "schedule",
                "csv": str(csv_path),
                "money_cols": ["amount"],
                "total_rows": [2],
            }
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    result = subprocess.run(
        ["python", "scripts/write_workbook.py", str(manifest_path)],
        cwd=".",
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, f"CLI failed: {result.stderr}"
    wb = load_workbook(out_path)
    assert wb.sheetnames == ["cover", "schedule"]
    assert wb["cover"]["A1"].value == "CLI test"
    assert wb["schedule"]["A1"].value == "account"
